import openpyxl as op
from os.path import dirname, join

path = join(dirname(__file__), "./FINAL450.xlsx")
wb = op.load_workbook(path).active

ls = []
for i in range(2, wb.max_row):
    k = []
    for j in range(1, 3):
        k.append(wb.cell(row=i, column=j).value)
    k.append(wb.cell(row=i, column=2).hyperlink.target)
    ls.append(k)

with open("./dsa450.json", "w") as f:
    f.write(str(ls))
