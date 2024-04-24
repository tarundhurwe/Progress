from openpyxl import load_workbook
from os.path import join, dirname

path = join(dirname(__file__), "Copy of Template Blind 75.xlsx")
workbook = load_workbook(path)

sheet = workbook.active

ls = []
count = 0
for row in sheet.iter_rows():
    if count != 0:
        ls.append([row[2].value, row[0].value, row[1].value, row[2].hyperlink.target])
    count += 1

workbook.close()

with open("problem.json", "w") as file:
    file.write(str(ls))
